from cers.hardware.models import Computer
from django.core.management import BaseCommand
from django.db import transaction
from openpyxl.reader.excel import load_workbook

EXCEL_MAPPING = {
    'name': 0,
    'model': 1,
    'mac_address': 2,
    'number': 3,
    'operating_system': 4,
    'office': 5,
    'bios_password': 6,
    'bitlocker': 7,
    'backup': 8,
    'ssd': 9,
    'norton': 10,
    'synology_pass': 11,
    'comment': 12,
    'vpn': 14,
    'serial_number': 15,
}

boolean_fields = ['backup', 'ssd', 'vpn']


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('--file')

    @transaction.atomic
    def handle(self, *args, **kwargs):
        if kwargs.get('file'):
            workbook = load_workbook(kwargs.get('file'), data_only=True)
            sheet = workbook.active
            for row_id, row in enumerate(sheet.iter_rows()):
                if row_id != 0:
                    fields = {}
                    for field_name, idx in EXCEL_MAPPING.items():
                        if row[idx]:
                            if field_name not in boolean_fields:
                                fields[field_name] = row[idx].value
                            elif row[idx].value:
                                fields[field_name] = True
                            else:
                                fields[field_name] = False

                    if fields != {}:
                        Computer.objects.create(**fields)
                    else:
                        break
