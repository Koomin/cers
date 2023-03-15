from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side


class Colors:
    GREY = 'C9C9C9'
    WHITE = 'FFFFFF'
    BLACK = '000000'


thin = Side(border_style='thin', color=Colors.BLACK)
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_borders(sheet, max_row, first_col=None, last_col=None):
    if not first_col and not last_col:
        first_col = sheet.min_column
        last_col = sheet.max_column
    for col in range(first_col, last_col + 1):
        for row in range(1, max_row + 1):
            sheet.cell(row, col).border = border


def color_row(sheet, row, color, first_col=None, last_col=None):
    color = Color(color)
    if not first_col and not last_col:
        first_col = sheet.min_column
        last_col = sheet.max_column
    for col in range(first_col, last_col + 1):
        sheet.cell(row, col).fill = PatternFill(patternType='solid', fgColor=color)


def color_font_row(sheet, row, color, first_col=None, last_col=None):
    if not first_col and not last_col:
        first_col = sheet.min_column
        last_col = sheet.max_column
    for col in range(first_col, last_col + 1):
        sheet.cell(row, col).font = Font(color=color)


def wrap_text(sheet, row, first_col, last_col):
    if not first_col and not last_col:
        first_col = sheet.min_column
        last_col = sheet.max_column
    for col in range(first_col, last_col):
        sheet.cell(row, col).alignment = Alignment(wrap_text=True, vertical='top')
